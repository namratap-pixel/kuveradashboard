"""
Parsers for the three ops-team Excel workbooks that back tabs 4, 5 and 6
of the dashboard: Attendance, Leave Management, Operations Roster.

These files are NOT live — they live in /data in the repo and are updated
by re-uploading to GitHub whenever HR/ops refreshes them. Render redeploys
automatically on push, so the dashboard picks up new data on the next deploy.
"""

import os
import re
import logging
from datetime import datetime, date

import openpyxl

logger = logging.getLogger(__name__)

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
ATTENDANCE_PATH = os.path.join(DATA_DIR, "Attendance.xlsx")
LEAVE_PATH = os.path.join(DATA_DIR, "Leave.xlsx")
ROSTER_PATH = os.path.join(DATA_DIR, "Roster.xlsx")

MONTH_ABBR = {
    1: ["jan"], 2: ["feb"], 3: ["mar"], 4: ["apr"], 5: ["may"], 6: ["jun"],
    7: ["jul", "july"], 8: ["aug"], 9: ["sep", "sept"], 10: ["oct"],
    11: ["nov"], 12: ["dec"],
}

# Attendance day-codes
PRESENT_CODES = {"P", "P_LL", "P_WH"}
LATE_LOGIN_CODES = {"P_LL"}
WFH_CODES = {"P_WH"}
PLANNED_LEAVE_CODES = {"PL"}
UNPLANNED_LEAVE_CODES = {"UL"}
WEEK_OFF_CODES = {"WO"}
OTHER_LEAVE_CODES = {"ML"}  # e.g. medical/maternity — not in the requested columns but not discarded


def _sheet_matches_month(title, month, year, exclude_chat=True):
    t = title.lower()
    if exclude_chat and "chat" in t:
        return False
    if str(year) not in t:
        return False
    return any(abbr in t for abbr in MONTH_ABBR[month])


def _find_sheet(wb, month, year, exclude_chat=True):
    for ws in wb.worksheets:
        if _sheet_matches_month(ws.title, month, year, exclude_chat):
            return ws
    return None


# ── Attendance ───────────────────────────────────────────────────────────

def get_attendance_month(year, month):
    """
    Returns {"days": [date,...], agents: {name: {"P":n,"P_LL":n,"P_WH":n,"PL":n,
    "UL":n,"WO":n,"ML":n,"present_days":n,"late_login":n,"wfh":n}}}
    """
    result = {"days": [], "agents": {}}
    if not os.path.exists(ATTENDANCE_PATH):
        return result
    try:
        wb = openpyxl.load_workbook(ATTENDANCE_PATH, data_only=True)
    except Exception as e:
        logger.error(f"Attendance workbook read error: {e}")
        return result

    ws = _find_sheet(wb, month, year)
    if ws is None:
        # "Current Month" sheet holds whatever month is currently in progress
        cm = wb["Current Month"] if "Current Month" in wb.sheetnames else None
        if cm is not None:
            hdr_month = cm.cell(row=1, column=9).value  # e.g. "July"
            hdr_year = cm.cell(row=1, column=11).value
            try:
                if hdr_year and int(hdr_year) == year and str(hdr_month).lower().startswith(MONTH_ABBR[month][0]):
                    ws = cm
            except Exception:
                pass
    if ws is None:
        return result

    # Find the "Name" label row; the row directly below it holds the actual dates
    label_row = None
    for r in range(1, 10):
        v = ws.cell(row=r, column=1).value
        if v and str(v).strip().lower() == "name":
            label_row = r
            break
    if label_row is None:
        return result
    header_row = label_row + 1

    date_cols = {}
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, datetime):
            date_cols[c] = v.date()

    result["days"] = sorted(set(date_cols.values()))

    name_row_start = header_row + 1
    for r in range(name_row_start, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if not name or not isinstance(name, str):
            continue
        name = name.strip()
        counts = {"P": 0, "P_LL": 0, "P_WH": 0, "PL": 0, "UL": 0, "WO": 0, "ML": 0}
        for c, d in date_cols.items():
            code = ws.cell(row=r, column=c).value
            if not code or not isinstance(code, str):
                continue
            code = code.strip().upper()
            if code in counts:
                counts[code] += 1
        counts["present_days"] = counts["P"] + counts["P_LL"] + counts["P_WH"]
        counts["late_login"] = counts["P_LL"]
        counts["wfh"] = counts["P_WH"]
        result["agents"][name] = counts

    return result


def get_todays_attendance_code(agent_name, on_date=None):
    """Quick lookup: what's this agent marked as today? Used for round-robin off-detection fallback."""
    on_date = on_date or date.today()
    data = get_attendance_month(on_date.year, on_date.month)
    if not data["agents"]:
        return None
    # We only stored aggregate counts above; re-read directly for the single day instead.
    if not os.path.exists(ATTENDANCE_PATH):
        return None
    try:
        wb = openpyxl.load_workbook(ATTENDANCE_PATH, data_only=True)
    except Exception:
        return None
    ws = _find_sheet(wb, on_date.month, on_date.year)
    if ws is None and "Current Month" in wb.sheetnames:
        ws = wb["Current Month"]
    if ws is None:
        return None
    label_row = None
    for r in range(1, 10):
        v = ws.cell(row=r, column=1).value
        if v and str(v).strip().lower() == "name":
            label_row = r
            break
    if label_row is None:
        return None
    header_row = label_row + 1
    target_col = None
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, datetime) and v.date() == on_date:
            target_col = c
            break
    if target_col is None:
        return None
    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if name and str(name).strip() == agent_name:
            code = ws.cell(row=r, column=target_col).value
            return str(code).strip().upper() if code else None
    return None


# ── Leave Management ─────────────────────────────────────────────────────

ANNUAL_LEAVE_ENTITLEMENT = 21
ANNUAL_OH_ENTITLEMENT = 4


def get_leave_balance():
    """Returns {name: {approved, unscheduled, lop, worked_on_mh, oh_available,
    oh_availed, pending_leaves, late_login, wfh, balance, oh_balance}}"""
    result = {}
    if not os.path.exists(LEAVE_PATH):
        return result
    try:
        wb = openpyxl.load_workbook(LEAVE_PATH, data_only=True)
        ws = wb["Summary"]
    except Exception as e:
        logger.error(f"Leave workbook read error: {e}")
        return result

    header_row = None
    for r in range(1, 5):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 12)]
        if vals[0] and str(vals[0]).strip().lower() == "name":
            header_row = r
            break
    if header_row is None:
        return result

    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if not name or not isinstance(name, str):
            continue

        def num(col):
            v = ws.cell(row=r, column=col).value
            try:
                return float(v)
            except (TypeError, ValueError):
                return 0.0

        approved = num(2)
        unscheduled = num(3)
        lop = num(4)
        worked_on_mh = num(5)
        oh_available = num(6)
        oh_availed = num(7)
        pending_leaves = num(9)
        late_login = num(10)
        wfh = num(12)

        used = approved + unscheduled
        result[name.strip()] = {
            "approved": approved,
            "unscheduled": unscheduled,
            "lop": lop,
            "worked_on_mh": worked_on_mh,
            "oh_available": oh_available,
            "oh_availed": oh_availed,
            "pending_leaves": pending_leaves,
            "late_login": late_login,
            "wfh": wfh,
            "annual_entitlement": ANNUAL_LEAVE_ENTITLEMENT,
            "used": used,
            "balance": round(ANNUAL_LEAVE_ENTITLEMENT - used, 1),
            "oh_entitlement": ANNUAL_OH_ENTITLEMENT,
            "oh_balance": round(ANNUAL_OH_ENTITLEMENT - oh_availed, 1),
        }
    return result


def get_holidays():
    """Returns list of {name, date, type} from Holiday List sheet."""
    holidays = []
    if not os.path.exists(LEAVE_PATH):
        return holidays
    try:
        wb = openpyxl.load_workbook(LEAVE_PATH, data_only=True)
        ws = wb["Holiday List"]
    except Exception:
        return holidays
    for row in ws.iter_rows(min_row=1, max_col=5):
        vals = [c.value for c in row]
        name, dt, htype = vals[1], vals[2], vals[3]
        if name and isinstance(dt, datetime):
            holidays.append({"name": name, "date": dt.date().isoformat(), "type": htype})
    return holidays


# ── Roster (for round-robin week-off detection) ─────────────────────────

def get_roster_status_for_date(on_date=None):
    """Returns {agent_name: 'WO'|'SHIFT'|None} — WO means on week off per roster,
    SHIFT means a shift time is listed (working day), None means not found in roster."""
    on_date = on_date or date.today()
    result = {}
    if not os.path.exists(ROSTER_PATH):
        return result
    try:
        wb = openpyxl.load_workbook(ROSTER_PATH, data_only=True)
    except Exception as e:
        logger.error(f"Roster workbook read error: {e}")
        return result

    ws = _find_sheet(wb, on_date.month, on_date.year, exclude_chat=True)
    if ws is None:
        return result

    # Header rows: row1 = weekday names, row2 = actual dates (may vary — scan first 3 rows)
    target_col = None
    date_header_row = None
    for r in range(1, 4):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, datetime) and v.date() == on_date:
                target_col = c
                date_header_row = r
                break
        if target_col:
            break
    if target_col is None:
        return result

    for r in range(date_header_row + 1, ws.max_row + 1):
        name = ws.cell(row=r, column=2).value  # "Employee Name" column
        if not name or not isinstance(name, str):
            continue
        name = name.strip()
        val = ws.cell(row=r, column=target_col).value
        if not val:
            continue
        val_s = str(val).strip().upper()
        if val_s == "WO":
            result[name] = "WO"
        elif re.match(r"^\d{1,2}:\d{2}\s*-\s*\d{1,2}:\d{2}$", val_s):
            result[name] = "SHIFT"
        else:
            result[name] = val_s  # e.g. PL, ML — treat as off/leave
    return result
